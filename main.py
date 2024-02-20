import pathlib
from tkcalendar import DateEntry
from tkinter import ttk
from openpyxl import load_workbook
import pandas as pd
from tkinter import *
from datetime import date
from word import GenerateWord
from tkinter import messagebox
from log import Log


class ExcelGUI:
    def __init__(self, root):
        try:
            self.window = root
            self.window.title("Aneliy App")
            self.window.geometry("850x480")
            self.window.config(bg="#CE7E50")
        except Exception as e:
            messagebox.showerror("ERROR", e)
            Log.add_new_error("", e)
            

        # Customization
        self.color_1 = "#828282"
        self.color_2 = "#828282"
        self.color_3 = "#66d9ff"
        self.color_4 = "black"
        self.color_5 = "#00ace6"
        self.font_1 = "orange"
        self.font_2 = "white"

        # Left Frame
        self.frame_1 = Frame(self.window, bg=self.color_1)
        self.frame_1.place(x=0, y=0, width=680, relheight=1)

        # Right Frame
        self.frame_2 = Frame(self.window, bg=self.color_2)
        self.frame_2.place(x=680, y=0, relwidth=1, relheight=1)

        self.name_var = StringVar()
        self.position_var = StringVar()
        self.education_var = StringVar()
        self.nkpd_var = StringVar()
        self.service_year_var = StringVar()
        self.service_month_var = StringVar()
        self.service_days_var = StringVar()
        self.professional_experience_days_var = StringVar()
        self.professional_experience_month_var = StringVar()
        self.professional_experience_year_var = StringVar()
        self.professional_experience_percent_var = StringVar()
        self.orz_var = StringVar()
        self.pks_lv_var = StringVar()
        self.brz_var = StringVar()
        self.signature_date_var = StringVar()
        self.education_degree_var = StringVar()
        self.education_speciality_var = StringVar()
        self.start_date_var = StringVar()
        self.salary_word_var = StringVar()
        self.free_food_var = StringVar(value="2.71")
        self.grounds_var = StringVar()
        self.contract_num_var = StringVar()
        self.professional_experience_money_var = StringVar()
        self.id_var = StringVar()
        self.address_var = StringVar()
        self.search_var = StringVar()
        self.order_num_var = StringVar()
        self.pedagog_var = StringVar()
        self.work_place_var = StringVar()
        self.type_var = StringVar()
        self.check_var = IntVar()
        self.var = []
        self.person_info = []
        self.work_place_values = ("ДГ“Калина Малина“ гр.Свищов, ул.“Хр.Павлович“",
                                  "ИГ в с.Българско Сливово към ДГ ”К. Малина” гр. Свищов",
                                  "ИГ в с.Вардим към  ДГ ”Калина Малина” гр. Свищов")
        self.type_values = ("Педагогически персонал",
                            "Помощен административен персонал",
                            "помощно-обсулжващ персонал")
        try:
            self.search()
            self.gui()
        except Exception as e:
            messagebox.showerror("ERROR", e)
            Log.add_new_error("", e)
            

    def gui(self):
        messagebox.showinfo("Напомняне", "Проверете дали данните в ексел са обновени и верни!")
        self.dropdown = ttk.Combobox(
            self.frame_2, textvariable=self.search_var, values=self.var, state="normal", width=30
        )
        self.dropdown.place(x=20, y=40, width=100)
        #Name
        self.name_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Имена")
        self.name_label.place(x=15, y=40, width=100)
        # self.name_label.pack()

        self.name_entry = Entry(self.frame_1, textvariable=self.name_var)
        self.name_entry.place(x=15, y=80, width=100)

        #position
        self.position_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5,text="Длъжност")
        self.position_label.place(x=125, y=40, width=100)

        self.position_entry = Entry(self.frame_1, textvariable=self.position_var)
        self.position_entry.place(x=125, y=80, width=100)

        #education
        self.education_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Образование")
        self.education_label.place(x=235, y=40, width=100)

        self.education_entry = Entry(self.frame_1, textvariable=self.education_degree_var)
        self.education_entry.place(x=235, y=80, width=100)

        #NKPD
        self.nkpd_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text=f"НКПД")
        self.nkpd_label.place(x=345, y=40, width=100)

        self.nkpd_entry = Entry(self.frame_1, textvariable=self.nkpd_var)
        self.nkpd_entry.place(x=345, y=80, width=100)

        #Service_year
        self.service_year_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Тр.стаж год.")
        self.service_year_label.place(x=455, y=40, width=100)

        self.service_year_entry = Entry(self.frame_1, textvariable=self.service_year_var)
        self.service_year_entry.place(x=455, y=80, width=100)

        # Service_month
        self.service_month_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Тр.стаж мес.")
        self.service_month_label.place(x=565, y=40, width=100)

        self.service_month_entry = Entry(self.frame_1, textvariable=self.service_month_var)
        self.service_month_entry.place(x=565, y=80, width=100)

        # Service_days
        self.service_days_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Тр.стаж дни")
        self.service_days_label.place(x=15, y=120, width=100)

        self.service_days_entry = Entry(self.frame_1, textvariable=self.service_days_var)
        self.service_days_entry.place(x=15, y=160, width=100)

        # professional_experience_years
        self.professional_experience_years_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5,
                                                         text="Проф.оп.год.")
        self.professional_experience_years_label.place(x=125, y=120, width=100)

        self.professional_experience_years_entry = Entry(self.frame_1, textvariable=self.professional_experience_year_var)
        self.professional_experience_years_entry.place(x=125, y=160, width=100)

            # professional_experience_month
        self.professional_experience_month_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5,
                                                         text="Проф.оп.мес.")
        self.professional_experience_month_label.place(x=235, y=120, width=100)

        self.professional_experience_month_entry = Entry(self.frame_1, textvariable=self.professional_experience_month_var)
        self.professional_experience_month_entry.place(x=235, y=160, width=100)

            # professional_experience_days
        self.professional_experience_days_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5,
                                                        text="Проф.оп.дни")
        self.professional_experience_days_label.place(x=345, y=120, width=100)

        self.professional_experience_days_entry = Entry(self.frame_1, textvariable=self.professional_experience_days_var)
        self.professional_experience_days_entry.place(x=345, y=160, width=100)

            # professional_experience_percent
        self.professional_experience_percent_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5,
                                                           text="% пр.вр.")
        self.professional_experience_percent_label.place(x=455, y=120, width=100)

        self.professional_experience_percent_entry = Entry(self.frame_1, textvariable=self.professional_experience_percent_var)
        self.professional_experience_percent_entry.place(x=455, y=160, width=100)

        # ORZ
        self.orz_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="ОРЗ")
        self.orz_label.place(x=565, y=120, width=100)

        self.orz_entry = Entry(self.frame_1, textvariable=self.orz_var)
        self.orz_entry.place(x=565, y=160, width=100)

        # professional_experience_money
        self.professional_experience_money_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5,
                                                         text="Проф. опит/лв.")
        self.professional_experience_money_label.place(x=15, y=200, width=100)

        self.professional_experience_money_entry = Entry(self.frame_1, textvariable=self.professional_experience_money_var)
        self.professional_experience_money_entry.place(x=15, y=240, width=100)

        # pks_lv
        self.pks_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="ПКС/лв.")
        self.pks_label.place(x=125, y=200, width=100)

        self.pks_entry = Entry(self.frame_1, textvariable=self.pks_lv_var)
        self.pks_entry.place(x=125, y=240, width=100)

        #contarct_num
        self.contract_num_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Труд. дог.-№")
        self.contract_num_label.place(x=235, y=200, width=100)

        self.contract_num_entry = Entry(self.frame_1, textvariable=self.contract_num_var)
        self.contract_num_entry.place(x=235, y=240, width=100)

        # ID
        self.id_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="ЕГН")
        self.id_label.place(x=345, y=200, width=100)

        self.id_entry = Entry(self.frame_1, textvariable=self.id_var)
        self.id_entry.place(x=345, y=240, width=100)

        # address
        self.address_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Адрес")
        self.address_label.place(x=455, y=200, width=100)

        self.address_entry = Entry(self.frame_1, textvariable=self.address_var)
        self.address_entry.place(x=455, y=240, width=100)

        # order_number
        self.grounds_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Основание")
        self.grounds_label.place(x=565, y=200, width=100)

        self.grounds_entry = Entry(self.frame_1, textvariable=self.grounds_var)
        self.grounds_entry.place(x=565, y=240, width=100)

        # signature_date #
        self.signature_date_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Дата-подп.")
        self.signature_date_label.place(x=15, y=280, width=100)

        self.signature_date_entry = DateEntry(self.frame_1, textvariable=self.signature_date_var)
        self.signature_date_entry.place(x=15, y=320, width=100)

        # education_degree
        self.education_degree_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="ПКС")
        self.education_degree_label.place(x=125, y=280, width=100)

        self.education_degree_entry = Entry(self.frame_1, textvariable=self.education_var)
        self.education_degree_entry.place(x=125, y=320, width=100)

        # education_speciality
        self.education_speciality_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Специалност")
        self.education_speciality_label.place(x=235, y=280, width=100)

        self.education_speciality_entry = Entry(self.frame_1, textvariable=self.education_speciality_var)
        self.education_speciality_entry.place(x=235, y=320, width=100)

        # start_date #
        self.start_date_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Считано от:")
        self.start_date_label.place(x=345, y=280, width=100)

        self.start_date_entry = DateEntry(self.frame_1, textvariable=self.start_date_var)
        self.start_date_entry.place(x=345, y=320, width=100)

        # salary_word
        self.salary_word_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5,
                                       text="ОВ с думи")
        self.salary_word_label.place(x=455, y=280, width=100)

        self.salary_word_entry = Entry(self.frame_1, textvariable=self.salary_word_var)
        self.salary_word_entry.place(x=455, y=320, width=100)

        # free_food
        self.free_food_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Безпл. храна")
        self.free_food_label.place(x=565, y=280, width=100)

        self.free_food_entry = Entry(self.frame_1, textvariable=self.free_food_var)
        self.free_food_entry.place(x=565, y=320, width=100)

        # order_num
        self.order_num_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Заповед №")
        self.order_num_label.place(x=15, y=360, width=100)
        #
        self.order_num_entry = Entry(self.frame_1, textvariable=self.order_num_var)
        self.order_num_entry.place(x=15, y=400, width=100)

        # педагог. стаж
        self.order_num_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5, text="Педагог. стаж")
        self.order_num_label.place(x=125, y=360, width=100)
        #
        self.order_num_entry = Entry(self.frame_1, textvariable=self.pedagog_var)
        self.order_num_entry.place(x=125, y=400, width=100)


        self.work_place_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5,
                                     text="Работно място")
        self.work_place_label.place(x=235, y=360, width=100)

        self.work_place_combbox = ttk.Combobox(self.frame_1, textvariable=self.work_place_var, values=self.work_place_values, state="normal")
        self.work_place_combbox.place(x=235, y=400, width=100)

        self.type_label = Label(self.frame_1, font=(self.font_2, 10, "bold"), bg=self.color_5,
                                      text="Кат. персонал")
        self.type_label.place(x=345, y=360, width=100)

        self.type_combbox = ttk.Combobox(self.frame_1, textvariable=self.type_var,
                                               values=self.type_values, state="normal")
        self.type_combbox.place(x=345, y=400, width=100)


        self.generate = Button(self.frame_1, text='Генерирай', font=(self.font_1, 12), bd=2,
                               cursor="hand2", bg=self.color_3, fg=self.color_4,
                               command=self.generate_btn).place(x=455, y=360, width=100)

        self.cancel = Button(self.frame_1, text='Изтрий', font=(self.font_1, 12), bd=2, command=self.cancel_btn,
                               cursor="hand2", bg=self.color_3, fg=self.color_4).place(x=565, y=360, width=100)

        show_info_button = Button(self.frame_2, text="Покажи информация", command=lambda: self.show_selected_info())
        show_info_button.place(x=15, y=80, width=120)
        chk = Checkbutton(self.frame_2, text='НПП', variable=self.check_var, bg="red", font=(self.font_2, 15, "bold"))
        chk.place(x=15, y=160, width=140)

    def search(self):
        self.excel_file = f"{pathlib.Path().resolve()}\data\shtatno-posledno.xlsx"
        self.workbook = load_workbook(self.excel_file)
        self.sheet = self.workbook.active
        try:
            row = 9
            while True:
                self.cell_values = [str(self.sheet.cell(row=row, column=3).value)]
                if self.cell_values[0] == "None":
                    break
                else:
                    self.var.extend(self.cell_values)
                    row += 1
            # for row_range in [(9, 39), (44, 51),(55, 57) ,(57, 66)]:
            #     for row in range(row_range[0], row_range[1] + 1):
            #         self.cell_values = [str(self.sheet.cell(row=row, column=3).value)]
            #         if self.cell_values[0].strip() not in ["СУМА ", "None", "Име и фамилия", "Н П П родени до 1960 год.",
            #                                        "П П родени след 60 год.", "П П - л родени след 1960 год.", "НПП", "СУМА", "Име и фамилия", "Име, презиме и фамилия"]:
            #             self.var.extend(self.cell_values)

        except Exception as e:
            messagebox.showerror("Error", e)
            Log.add_new_error("", e)
            

    def fill_entry(self):
        try:
            self.name_var.set(self.person_info[0].get())
            self.position_var.set(self.person_info[1].get())
            self.education_var.set(self.person_info[2].get())
            self.nkpd_var.set(self.person_info[3].get())
            self.service_year_var.set(self.person_info[4].get())
            self.service_month_var.set(self.person_info[5].get())
            self.service_days_var.set(self.person_info[6].get())
            self.professional_experience_days_var.set(self.person_info[9].get())
            self.professional_experience_month_var.set(self.person_info[8].get())
            self.professional_experience_year_var.set(self.person_info[7].get())
            self.professional_experience_percent_var.set(self.person_info[10].get())
            self.orz_var.set(self.person_info[11].get())
            self.pks_lv_var.set(self.person_info[14].get())
            self.brz_var.set(self.person_info[15].get())
            self.signature_date_var.set("")
            self.education_degree_var.set(self.person_info[19].get())
            self.pedagog_var.set(self.person_info[20].get())
            self.education_speciality_var.set(self.person_info[21].get())
            self.start_date_var.set("")
            self.salary_word_var.set("")
            self.free_food_var.set(value="2.71")
            self.grounds_var.set("")
            self.order_num_var.set("")
            self.contract_num_var.set(self.person_info[18].get())
            self.professional_experience_money_var.set(f"{float(self.person_info[12].get()):.2f}")
            self.id_var.set(self.person_info[16].get())
            self.address_var.set(self.person_info[17].get())
            self.check_var.set(0)
        except Exception as e:
            messagebox.showerror("Грешка", e)
            Log.add_new_error("", e)


    def show_selected_info(self):
        self.selected_name = self.search_var.get()
        if self.selected_name:
            row = 9
            while True:
                cell_value = self.sheet.cell(row=row, column=3).value
                if cell_value == "None":
                    break
                elif self.selected_name.lower() in str(cell_value).lower():
                    data = pd.read_excel(f"{pathlib.Path().resolve()}\data\\shtatno-posledno.xlsx")

                    # Вземане на информация от ред 9 в колоните от "C" до "R"
                    row_9_data = data.iloc[row-2, 2:25]

                    # Създаване на текстови полета за показване на информацията
                    for col, value in zip(row_9_data.index, row_9_data.values):
                        self.person_info.append(StringVar(value=value))
                    break
                row += 1
            self.fill_entry()
            # # Извличане на информацията за избрания човек от Excel таблицата
            # for row_range in [(9, 39), (44, 51),(55, 57),(57, 66)]:
            #     for row in range(row_range[0], row_range[1] + 1):
            #         cell_value = self.sheet.cell(row=row, column=3).value
            #         if self.selected_name.lower() in str(cell_value).lower():
            #             data = pd.read_excel(f"{pathlib.Path().resolve()}\data\\shtatno-posledno.xlsx")
            #
            #             # Вземане на информация от ред 9 в колоните от "C" до "R"
            #             row_9_data = data.iloc[row-2, 2:25]
            #
            #             # Създаване на текстови полета за показване на информацията
            #             for col, value in zip(row_9_data.index, row_9_data.values):
            #                 self.person_info.append(StringVar(value=value))
        self.fill_entry()

    def cancel_btn(self):
        self.name_var.set("")
        self.position_var.set("")
        self.education_var.set("")
        self.nkpd_var.set("")
        self.service_year_var.set("")
        self.service_month_var.set("")
        self.service_days_var.set("")
        self.professional_experience_days_var.set("")
        self.professional_experience_month_var.set("")
        self.professional_experience_year_var.set("")
        self.professional_experience_percent_var.set("")
        self.orz_var.set("")
        self.pks_lv_var.set("")
        self.brz_var.set("")
        self.signature_date_var.set("")
        self.education_degree_var.set("")
        self.education_speciality_var.set("")
        self.start_date_var.set("")
        self.salary_word_var.set("")
        self.free_food_var.set(value="2.71")
        self.grounds_var.set("")
        self.contract_num_var.set("")
        self.professional_experience_money_var.set("")
        self.id_var.set("")
        self.address_var.set("")
        self.search_var.set("")
        self.order_num_var.set("")
        self.pedagog_var.set("")
        self.check_var.set(0)
        self.var = []
        self.person_info = []

    def generate_btn(self):
        messagebox.showwarning("showwarning", "Проверете данните дали са правилно въведени!")
        test = self.pedagog_var.get().split("/")
        data = self.signature_date_var.get().split("/")
        f_l_name = self.name_var.get().split()
        try:
            self.replacements = {
                '<<id>>': self.id_var.get(),
                '<<signature_date>>': str(self.signature_date_entry.get_date().strftime("%d.%m.%Y")),
                "<<grounds>>": self.grounds_var.get(),
                "<<name>>": self.name_var.get(),
                "<<address>>": self.address_var.get(),
                "<<education_degree>>": self.education_degree_var.get(),
                "<<education_speciality>>": self.education_speciality_var.get(),
                "<<service_date>>": str(self.service_year_var.get() + "г. " + self.service_month_var.get() + "м. " + self.service_days_var.get() + "д."),
                "<<professional_date>>": str(self.professional_experience_year_var.get() + "г. " +
                                             self.professional_experience_month_var.get() + "м. " + self.professional_experience_days_var.get() + "д."),
                "<<order_num>>": self.order_num_var.get(),
                "<<position>>": self.position_var.get(),
                "<<NKPD>>": self.nkpd_var.get(),
                "<<ProsVr>>": f"{float(self.professional_experience_percent_var.get())*100:.0f}",
                "<<sum_pros>>": self.professional_experience_money_var.get(),
                "<<free_food>>": self.free_food_var.get(),
                "<<BMS>>": self.orz_var.get(),
                "<<contract_num>>": self.contract_num_var.get(),
                "<<salary_word>>": self.salary_word_var.get(),
                "<<start_date>>": str(self.start_date_entry.get_date().strftime("%d.%m.%Y.")),
                "<<pks_lv>>": self.pks_lv_var.get(),
                "<<pks>>": self.education_var.get(),
                "<<category>>": self.type_var.get(),
                "<<work_place>>": self.work_place_var.get(),
                "<<f_l_name>>": str(f_l_name[0] + " " + f_l_name[2])

            }
        except Exception as e:
            messagebox.showerror("Грешка", e)
            Log.add_new_error("", e)
            


        if self.check_var.get() == 1:
            try:
                GenerateWord().replace_text_in_docx(self.replacements, f"{pathlib.Path().resolve()}\data\\temp.docx")
            except Exception as e:
                messagebox.showerror("Грешка", e)
                Log.add_new_error("", e)
                

        else:
            try:
                self.replacements["<<pedagog>>"] = str(test[0]+"г. "+test[1]+"м. "+test[2]+"д")
                GenerateWord().replace_text_in_docx(self.replacements, f"{pathlib.Path().resolve()}\data\\temp_pp.docx")
            except Exception as e:
                messagebox.showerror("Грешка", e)
                Log.add_new_error("", e)
                


if __name__ == "__main__":
    try:
        root = Tk()
        excel_gui = ExcelGUI(root)
        root.mainloop()
    except Exception as e:
        messagebox.showerror("Greshka", e)
        Log.add_new_error("", e)
        
